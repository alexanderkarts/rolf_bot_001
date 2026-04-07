import logging
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, ContextTypes, MessageHandler, filters
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from config import TELEGRAM_TOKEN, SHEET_URL, GOOGLE_CREDENTIALS_JSON, GOOGLE_SHEET_URL
from utils import full_stock, PHOTO_COLUMNS

logging.basicConfig(format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

reply_keyboard = [
    ["Статистика"],
    ["Полный сток", "Авто без фото"],
    ["Авто без места хранения"],
    ["Переданные авто", "Не переданные авто"],
    ["Поиск ключа"]
]
markup = ReplyKeyboardMarkup(reply_keyboard, resize_keyboard=True)

scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
creds = ServiceAccountCredentials.from_json_keyfile_name(GOOGLE_CREDENTIALS_JSON, scope)
client = gspread.authorize(creds)
gs_sheet = client.open_by_url(GOOGLE_SHEET_URL).sheet1

WAITING_KEY = {}
keys_data = {}  # ключи и их статусы: sent / not_sent

IGNORE_WORDS = ["ключ", "ключа", "ключей", "кл"]

def extract_key_number(text: str):
    if not text:
        return None
    text = text.lower()
    text = re.sub(r'\d+\s*(ключ|ключа|ключей|кл)', '', text)
    numbers = re.findall(r'\d+', text)
    if not numbers:
        return None
    return int(numbers[0])

def load_data() -> pd.DataFrame:
    try:
        df = pd.read_csv(SHEET_URL)
        logger.info("Данные загружены")
        return df
    except Exception as e:
        logger.error(f"Ошибка загрузки данных: {e}")
        return pd.DataFrame()

def format_excel(file_path: str):
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                value = str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2
    wb.save(file_path)

def write_df_to_sheet(sheet, df: pd.DataFrame):
    df_clean = df.fillna("")
    values = [df_clean.columns.tolist()] + df_clean.values.tolist()
    sheet.clear()
    sheet.update(range_name="A1", values=values)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Привет! Выберите действие:", reply_markup=markup)

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global keys_data
    message = update.message
    text = message.text or message.caption
    chat_type = message.chat.type

    df = load_data()
    if df.empty:
        await message.reply_text("Ошибка: нет данных для отображения.")
        return

    # ===== Проверка пересылаемых сообщений в группе =====
    if chat_type in ["group", "supergroup"]:
        key_number = extract_key_number(text)
        df_photo_keys = df[df["Кол-во фото для сайта"] == 0]["Номер ключа"].astype(str).tolist()
        if key_number and str(key_number) in df_photo_keys:
            keys_data[str(key_number)] = {"VIN": df[df["Номер ключа"] == key_number]["VIN"].values[0], "status": "sent"}
            await message.reply_text(f"✅ Записан ключ {key_number} как переданный")
            logger.info(f"Ключ {key_number} записан как переданный")
            return

    # ===== Обработка кнопок и поиска ключа =====
    if text == "Полный сток":
        df_full = full_stock(df)
        file_path = "full_stock.xlsx"
        df_full.to_excel(file_path, index=False)
        format_excel(file_path)
        write_df_to_sheet(gs_sheet, df_full)
        await message.reply_text(f"Полный сток готов. Всего машин: {len(df_full)}")
        await message.reply_document(open(file_path, "rb"))
        return

    elif text == "Авто без фото":
        df_photo = df[df["Кол-во фото для сайта"] == 0]
        df_photo = df_photo.sort_values(by="Дней с даты поступления", ascending=False)
        df_photo = df_photo[PHOTO_COLUMNS]
        file_path = "auto_without_photo.xlsx"
        df_photo.to_excel(file_path, index=False)
        format_excel(file_path)
        write_df_to_sheet(gs_sheet, df_photo)
        await message.reply_text(f"Авто без фото готово. Всего машин: {len(df_photo)}")
        await message.reply_document(open(file_path, "rb"))
        return

    elif text == "Авто без места хранения":
        df_full = full_stock(df)
        df_no_storage = df_full[
            (df_full["Место хранения"].isna() | (df_full["Место хранения"] == "")) &
            (df_full["Кол-во фото для сайта"] != 0)
        ]
        df_no_storage = df_no_storage.sort_values(by="Дней с даты поступления", ascending=True)
        file_path = "Авто_без_места.xlsx"
        df_no_storage.to_excel(file_path, index=False)
        format_excel(file_path)
        write_df_to_sheet(gs_sheet, df_no_storage)
        await message.reply_document(open(file_path, "rb"), caption=f"Авто без места хранения: {len(df_no_storage)} машин")
        return

    elif text == "Статистика":
        df_full = full_stock(df)
        total_stock = len(df_full)
        no_photo = len(df_full[df_full["Кол-во фото для сайта"] == 0])
        no_storage = len(df_full[
            (df_full["Место хранения"].isna() | (df_full["Место хранения"] == "")) &
            (df_full["Кол-во фото для сайта"] != 0)
        ])
        parking_counts = df_full["Место хранения"].fillna("Без места").value_counts()
        msg = f"📊 Статистика\n\nПолный сток: {total_stock}\nАвто без фото: {no_photo}\nАвто без места хранения: {no_storage}\n\nПарковки:\n"
        for parking, count in parking_counts.items():
            msg += f"{parking}: {count}\n"
        await message.reply_text(msg)
        return

    elif text == "Переданные авто":
        df_photo = df[df["Кол-во фото для сайта"] == 0]
        sent_keys = {k:v for k,v in keys_data.items() if v.get("status") == "sent" and k in df_photo["Номер ключа"].astype(str).tolist()}
        if not sent_keys:
            await message.reply_text("Нет переданных авто.")
            return
        df_sent = pd.DataFrame([{"Номер ключа": k, "VIN": v.get("VIN")} for k,v in sent_keys.items()])
        file_path = "Переданные_авто.xlsx"
        df_sent.to_excel(file_path, index=False)
        format_excel(file_path)
        write_df_to_sheet(gs_sheet, df_sent)
        await message.reply_document(open(file_path, "rb"))
        return

    elif text == "Не переданные авто":
        df_photo = df[df["Кол-во фото для сайта"] == 0]
        not_sent_keys = df_photo[~df_photo["Номер ключа"].astype(str).isin(keys_data.keys())]
        file_path = "Не_переданные_авто.xlsx"
        not_sent_keys.to_excel(file_path, index=False)
        format_excel(file_path)
        write_df_to_sheet(gs_sheet, not_sent_keys)
        await message.reply_document(open(file_path, "rb"))
        return

    elif text == "Поиск ключа":
        WAITING_KEY[message.chat_id] = True
        await message.reply_text("Введите номер ключа")
        return

    elif WAITING_KEY.get(message.chat_id):
        WAITING_KEY.pop(message.chat_id)
        if not text.isdigit():
            await message.reply_text("Номер ключа должен быть числом")
            return
        key_number = int(text)
        df_full = full_stock(df)
        car = df_full[df_full["Номер ключа"] == key_number]
        if car.empty:
            await message.reply_text("Машина не найдена")
            return
        car = car.iloc[0]
        msg = (
            f"🚗 Автомобиль найден\n\n"
            f"Номер ключа: {car['Номер ключа']}\n"
            f"Марка: {car['Марка']}\n"
            f"Модель: {car['Модель']}\n"
            f"VIN: {car['VIN']}\n"
            f"Год выпуска: {car['Год выпуска']}\n"
            f"Пробег: {car['Пробег']}\n"
            f"Цвет: {car['Цветкузова']}\n"
            f"Рег. номер: {car['Рег. номер']}\n"
            f"Парковка: {car['Место хранения']}\n"
            f"Дней в стоке: {car['Дней с даты поступления']}\n"
            f"Цена продажи: {car['Цена продажи']}\n"
            f"Байер: {car['Байер']}\n"
            f"Тип сделки: {car['Тип сделки']}"
        )
        await message.reply_text(msg)
        return

    await message.reply_text("Неизвестная команда. Попробуйте ещё раз.")

if __name__ == "__main__":
    app = ApplicationBuilder().token(TELEGRAM_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.ALL, handle_message))  # Ловим все сообщения
    logger.info("Бот запущен")
    app.run_polling()